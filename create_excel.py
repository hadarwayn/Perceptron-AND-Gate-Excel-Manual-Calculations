#!/usr/bin/env python3
"""
Create a FUN, colorful Excel file that teaches Perceptron Learning!
Designed so a 15-year-old can understand every single step.

Usage: python create_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# Color scheme - matching PRD specification
COLORS = {
    'blue_header': 'FF4472C4',
    'yellow_weights': 'FFFFF2CC',
    'purple_inputs': 'FFE2D0F8',
    'green_correct': 'FFC6EFCE',
    'red_incorrect': 'FFFFC7CE',
    'light_blue_docs': 'FFDEEBF7',
    'orange_calc': 'FFFCE4D6',
    'white': 'FFFFFFFF',
}


def create_perceptron_excel():
    """Create the educational Perceptron AND Gate Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Perceptron Learning"

    # Set column widths for readability
    col_widths = {'A': 8, 'B': 8, 'C': 6, 'D': 6, 'E': 6, 'F': 8, 'G': 8, 'H': 8,
                  'I': 8, 'J': 12, 'K': 10, 'L': 8, 'M': 12, 'N': 10, 'O': 10, 'P': 10, 'Q': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # === SECTION 1: Welcome Header (Row 1-2) ===
    ws.merge_cells('A1:Q1')
    ws['A1'] = "PERCEPTRON LEARNING - Watch AI Learn Step by Step!"
    ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORS['blue_header'], fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:Q2')
    ws['A2'] = "This Excel shows how a simple 'brain' learns the AND gate. Change the YELLOW cells to experiment!"
    ws['A2'].font = Font(size=11, italic=True)
    ws['A2'].fill = PatternFill(start_color=COLORS['light_blue_docs'], fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center')

    # === SECTION 2: AND Truth Table (Row 4-9) ===
    ws['A4'] = "THE AND GATE TRUTH TABLE"
    ws['A4'].font = Font(size=14, bold=True)
    ws.merge_cells('A4:F4')

    headers = ['x0 (bias)', 'x1', 'x2', 'Output', 'Plain English']
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=i+1, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['blue_header'], fill_type='solid')
        cell.border = thin_border

    # Truth table data: x0=1 (bias), x1, x2, output, explanation
    truth_table = [
        [1, 0, 0, 0, "OFF + OFF = OFF"],
        [1, 0, 1, 0, "OFF + ON = OFF"],
        [1, 1, 0, 0, "ON + OFF = OFF"],
        [1, 1, 1, 1, "ON + ON = ON!"],
    ]
    for row_idx, row_data in enumerate(truth_table, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx <= 3:  # Input columns
                cell.fill = PatternFill(start_color=COLORS['purple_inputs'], fill_type='solid')

    # === SECTION 3: Initial Weights (Row 11-15) ===
    ws['A11'] = "INITIAL WEIGHTS (The Brain's Starting Guesses)"
    ws['A11'].font = Font(size=14, bold=True)
    ws.merge_cells('A11:E11')

    ws['A12'] = "CHANGE THESE YELLOW CELLS TO EXPERIMENT!"
    ws['A12'].font = Font(bold=True, color='FF0000')
    ws.merge_cells('A12:E12')

    weights_info = [
        ('W0 (Bias Weight):', 'C13', 3, "Controls the 'default' guess"),
        ('W1 (x1 Weight):', 'C14', 3, "How important is x1?"),
        ('W2 (x2 Weight):', 'C15', 3, "How important is x2?"),
    ]
    for i, (label, cell_ref, value, desc) in enumerate(weights_info):
        ws[f'A{13+i}'] = label
        ws[cell_ref] = value
        ws[cell_ref].fill = PatternFill(start_color=COLORS['yellow_weights'], fill_type='solid')
        ws[cell_ref].font = Font(bold=True, size=12)
        ws[cell_ref].border = thin_border
        ws[f'D{13+i}'] = desc
        ws[f'D{13+i}'].font = Font(italic=True, color='666666')

    # === SECTION 4: Learning Table Headers (Row 17-19) ===
    ws['A17'] = "WATCH THE PERCEPTRON LEARN - Every Calculation Visible!"
    ws['A17'].font = Font(size=14, bold=True)
    ws.merge_cells('A17:Q17')

    # Group headers
    group_headers = [('A18', 'A18', '#', COLORS['white']),
                     ('B18', 'B18', 'Sample', COLORS['white']),
                     ('C18', 'E18', 'INPUTS', COLORS['purple_inputs']),
                     ('F18', 'F18', 'Target', COLORS['white']),
                     ('G18', 'I18', 'CURRENT WEIGHTS', COLORS['yellow_weights']),
                     ('J18', 'J18', 'Z (Dot)', COLORS['orange_calc']),
                     ('K18', 'K18', 'Pred', COLORS['white']),
                     ('L18', 'L18', 'Error', COLORS['white']),
                     ('M18', 'M18', 'Status', COLORS['white']),
                     ('N18', 'P18', 'NEW WEIGHTS', COLORS['yellow_weights']),
                     ('Q18', 'Q18', 'Result', COLORS['white'])]

    for start, end, text, color in group_headers:
        if start != end:
            ws.merge_cells(f'{start}:{end}')
        ws[start] = text
        ws[start].font = Font(bold=True)
        ws[start].fill = PatternFill(start_color=color, fill_type='solid')
        ws[start].alignment = Alignment(horizontal='center')
        ws[start].border = thin_border

    # Column headers row 19
    col_headers = ['#', 'Samp', 'x0', 'x1', 'x2', 'Actual', 'W0', 'W1', 'W2',
                   'Z=W*X', 'Pred', 'Err', 'Status', 'W0\'', 'W1\'', 'W2\'', '']
    for i, h in enumerate(col_headers):
        cell = ws.cell(row=19, column=i+1, value=h)
        cell.font = Font(bold=True, size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # === SECTION 5: 20 Learning Iterations with Formulas (Row 20-39) ===
    for iteration in range(1, 21):
        row = 19 + iteration
        sample = ((iteration - 1) % 4) + 1  # Cycles 1,2,3,4,1,2,3,4...

        ws.cell(row=row, column=1, value=iteration).border = thin_border
        ws.cell(row=row, column=2, value=sample).border = thin_border

        # x0, x1, x2 from truth table using INDEX
        ws.cell(row=row, column=3, value=f'=INDEX($A$6:$A$9,B{row})').border = thin_border
        ws.cell(row=row, column=4, value=f'=INDEX($B$6:$B$9,B{row})').border = thin_border
        ws.cell(row=row, column=5, value=f'=INDEX($C$6:$C$9,B{row})').border = thin_border
        # Actual output from truth table
        ws.cell(row=row, column=6, value=f'=INDEX($D$6:$D$9,B{row})').border = thin_border

        # Current weights (from initial or previous row's new weights)
        if iteration == 1:
            ws.cell(row=row, column=7, value='=$C$13').border = thin_border
            ws.cell(row=row, column=8, value='=$C$14').border = thin_border
            ws.cell(row=row, column=9, value='=$C$15').border = thin_border
        else:
            ws.cell(row=row, column=7, value=f'=N{row-1}').border = thin_border
            ws.cell(row=row, column=8, value=f'=O{row-1}').border = thin_border
            ws.cell(row=row, column=9, value=f'=P{row-1}').border = thin_border

        # Z = dot product = W0*x0 + W1*x1 + W2*x2
        ws.cell(row=row, column=10, value=f'=G{row}*C{row}+H{row}*D{row}+I{row}*E{row}').border = thin_border
        # Predicted = step function
        ws.cell(row=row, column=11, value=f'=IF(J{row}>0,1,0)').border = thin_border
        # Error = Actual - Predicted
        ws.cell(row=row, column=12, value=f'=F{row}-K{row}').border = thin_border
        # Status text
        ws.cell(row=row, column=13, value=f'=IF(L{row}=0,"CORRECT","WRONG")').border = thin_border
        # New weights: W_new = W_old + error * x
        ws.cell(row=row, column=14, value=f'=G{row}+L{row}*C{row}').border = thin_border
        ws.cell(row=row, column=15, value=f'=H{row}+L{row}*D{row}').border = thin_border
        ws.cell(row=row, column=16, value=f'=I{row}+L{row}*E{row}').border = thin_border
        # Visual result
        ws.cell(row=row, column=17, value=f'=IF(L{row}=0,"V","X")').border = thin_border

        # Apply colors to columns
        for col in [3, 4, 5]:  # Inputs - purple
            ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['purple_inputs'], fill_type='solid')
        for col in [7, 8, 9, 14, 15, 16]:  # Weights - yellow
            ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['yellow_weights'], fill_type='solid')
        ws.cell(row=row, column=10).fill = PatternFill(start_color=COLORS['orange_calc'], fill_type='solid')

    # === SECTION 6: Summary (Row 41-46) ===
    ws['A41'] = "LEARNING SUMMARY - How Did Our Robot Do?"
    ws['A41'].font = Font(size=14, bold=True)
    ws.merge_cells('A41:F41')

    summary = [
        ('Total Iterations:', '=20', 'A42', 'B42'),
        ('Correct:', '=COUNTIF(M20:M39,"CORRECT")', 'A43', 'B43'),
        ('Wrong:', '=COUNTIF(M20:M39,"WRONG")', 'A44', 'B44'),
        ('Final W0:', '=N39', 'A45', 'B45'),
        ('Final W1:', '=O39', 'C45', 'D45'),
        ('Final W2:', '=P39', 'E45', 'F45'),
    ]
    for label, formula, label_cell, value_cell in summary:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = formula
        ws[value_cell].border = thin_border

    ws['B43'].fill = PatternFill(start_color=COLORS['green_correct'], fill_type='solid')
    ws['B44'].fill = PatternFill(start_color=COLORS['red_incorrect'], fill_type='solid')

    # Save the file
    output_path = Path(__file__).parent / "results" / "Perceptron_AND_Gate_Learning.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_perceptron_excel()
